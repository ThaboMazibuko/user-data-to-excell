import openpyxl
from openpyxl.utils import get_column_letter

class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename
        self.wb = self.load_workbook()

    def load_workbook(self):
        try:
            return openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            return openpyxl.Workbook()

    def save_to_excel(self, data):
        sheet = self.wb.active
        sheet.append(data)
        self.wb.save(self.filename)

    def read_from_excel(self):
        sheet = self.wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data