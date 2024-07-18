import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.utils import get_column_letter

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        self.name_label = tk.Label(self, text="Name:", font=("Helvetica", 14), fg="blue")
        self.name_label.grid(column=0, row=0, padx=10, pady=10)
        self.name_entry = tk.Entry(self, width=30, font=("Helvetica", 14))
        self.name_entry.grid(column=1, row=0, padx=10, pady=10)

        self.surname_label = tk.Label(self, text="Surname:", font=("Helvetica", 14), fg="blue")
        self.surname_label.grid(column=0, row=1, padx=10, pady=10)
        self.surname_entry = tk.Entry(self, width=30, font=("Helvetica", 14))
        self.surname_entry.grid(column=1, row=1, padx=10, pady=10)

        self.email_label = tk.Label(self, text="Email:", font=("Helvetica", 14), fg="blue")
        self.email_label.grid(column=0, row=2, padx=10, pady=10)
        self.email_entry = tk.Entry(self, width=30, font=("Helvetica", 14))
        self.email_entry.grid(column=1, row=2, padx=10, pady=10)

        self.cellphone_label = tk.Label(self, text="Cellphone Number:", font=("Helvetica", 14), fg="blue")
        self.cellphone_label.grid(column=0, row=3, padx=10, pady=10)
        self.cellphone_entry = tk.Entry(self, width=30, font=("Helvetica", 14))
        self.cellphone_entry.grid(column=1, row=3, padx=10, pady=10)

        self.save_button = tk.Button(self, text="Save", command=self.save_to_excel, font=("Helvetica", 16), fg="green", bg="light green")
        self.save_button.grid(column=1, row=4, padx=10, pady=10)

        self.data_text = tk.Text(self, width=40, height=10, font=("Helvetica", 12))
        self.data_text.grid(column=0, row=5, columnspan=2, padx=10, pady=10)

    def save_to_excel(self):
        try:
            # Create a new Excel file if it doesn't exist
            try:
                wb = openpyxl.load_workbook("user_data.xlsx")
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            # Get the active sheet
            sheet = wb.active

            # Get the data from the entry fields
            name = self.name_entry.get()
            surname = self.surname_entry.get()
            email = self.email_entry.get()
            cellphone = self.cellphone_entry.get()

            # Add the data to the Excel file
            sheet.append([name, surname, email, cellphone])

            # Save the changes
            wb.save("user_data.xlsx")

            # Clear the entry fields
            self.name_entry.delete(0, tk.END)
            self.surname_entry.delete(0, tk.END)
            self.email_entry.delete(0, tk.END)
            self.cellphone_entry.delete(0, tk.END)

            # Show a confirmation message
            messagebox.showinfo("Success", "Data saved ")

           # Update the Text widget with the data from the Excel file
            self.data_text.delete(1.0, tk.END)  # Clear the Text widget

            # Add headings
            headings = ["Name", "Surname", "Email", "Cellphone Number"]
            self.data_text.tag_config("heading", font=("Helvetica", 12, "bold"), foreground="blue")
            self.data_text.insert(tk.END, "\t".join(headings) + "\n", "heading")

            # Add data
            for row in sheet.iter_rows(values_only=True):
                self.data_text.insert(tk.END, "\t".join(map(str, row)) + "\n")
        except Exception as e:
            messagebox.showerror("Error", str(e))

root = tk.Tk()
root.state("zoomed")  # Maximize the GUI on runtime
app = Application(master=root)
app.mainloop()